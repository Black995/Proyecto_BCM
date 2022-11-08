import csv
import os
from django.conf import settings
from bcm_phase2.models import (R_SO, InterestedParty, Ressource, ServiceOffered, ServiceUsed, Staff, 
                                OrganizationActivity, SO_S)
from .serializers import (InterestedPartyListSerializer, OrganizationActivityListSerializer, OrganizationActivitySerializer, R_SOSerializer, RessourceListSerializer, RessourceSerializer, RessourceWithServiceOfferedSerializer, 
                            ServiceOfferedListSerializer, ServiceOfferedSerializer, ServiceUsedListSerializer, 
                            ServiceUsedSerializer, StaffListSerializer, StaffSerializer, interestedPartySerializer,
                            SO_SSerializer, ServiceOfferedWithStaffsSerializer)
from rest_framework import serializers, status, viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Q, F
from bcm_phase2.api.filters import (R_SOFilterBackend, SO_SFilterBackend)
from rest_framework.generics import (ListAPIView)
from django.http.response import Http404
from django.shortcuts import HttpResponse
from .utils import render_to_pdf

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction
from django.shortcuts import get_object_or_404
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.decorators import action


class ServiceOfferedListViewSet(viewsets.ModelViewSet):
    model = ServiceOffered
    queryset = ServiceOffered.objects.all().order_by('name')
    serializer_class = ServiceOfferedListSerializer


class ServiceOfferedViewSet(viewsets.ModelViewSet):
    model = ServiceOffered
    queryset = ServiceOffered.objects.all()
    serializer_class = ServiceOfferedSerializer


class ServiceOfferedWithStaffsViewSet(viewsets.ModelViewSet):
    model = ServiceOffered
    queryset = ServiceOffered.objects.all()
    serializer_class = ServiceOfferedWithStaffsSerializer



class ServiceUsedListViewSet(viewsets.ModelViewSet):
    model = ServiceUsed
    queryset = ServiceUsed.objects.all().order_by('name')
    serializer_class = ServiceUsedListSerializer


class ServiceUsedViewSet(viewsets.ModelViewSet):
    model = ServiceUsed
    queryset = ServiceUsed.objects.all()
    serializer_class = ServiceUsedSerializer


class StaffListViewSet(viewsets.ModelViewSet):
    model = Staff
    queryset = Staff.objects.all().order_by('staff_number')
    serializer_class = StaffListSerializer


class StaffViewSet(viewsets.ModelViewSet):
    model = Staff
    queryset = Staff.objects.all()
    serializer_class = StaffSerializer

class OrganizationActivityListViewSet(viewsets.ModelViewSet):
    model = OrganizationActivity
    queryset = OrganizationActivity.objects.all().order_by('name')
    serializer_class = OrganizationActivityListSerializer

class OrganizationActivityViewSet(viewsets.ModelViewSet):
    model = OrganizationActivity
    queryset = OrganizationActivity.objects.all()

    serializer_class = OrganizationActivitySerializer


class InterestedPartyListViewSet(viewsets.ModelViewSet):
    model = InterestedParty
    queryset = InterestedParty.objects.all().order_by('name')
    serializer_class = InterestedPartyListSerializer


class InterestedPartyViewSet(viewsets.ModelViewSet):
    model = InterestedParty
    queryset = InterestedParty.objects.all()
    serializer_class = interestedPartySerializer


class SO_SViewSet(viewsets.ModelViewSet):
    model = SO_S
    queryset = SO_S.objects.all()
    serializer_class = SO_SSerializer
    filter_backends = [SO_SFilterBackend, ]

    

class download_bulk_upload_template(ListAPIView):
    queryset = Staff.objects.all()

    def get(self, request):
        file_path = os.path.join(
            settings.STATICFILES_DIRS[2], 'bulk_upload_staff.xlsx')
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(
                    fh.read(), content_type="application/vnd.ms-Excel")
                response['Content-Disposition'] = 'inline; filename=' + \
                    os.path.basename(file_path)
                return response
        raise Http404


class BulkUploadViewSet(viewsets.ModelViewSet):
    bulk_upload_template_name = None

    @staticmethod
    def clear_worksheet_styles_and_comments(worksheet: Worksheet):
        is_first_row = True
        for row in worksheet.rows:
            if is_first_row:
                is_first_row = False
                continue
            for cell in row:
                # Removing comments and background colors
                cell.comment = None
                cell.fill = openpyxl.styles.PatternFill(fill_type=None)

    @action(methods=['POST'], detail=False)
    @transaction.non_atomic_requests()
    def bulk_upload(self, request):
        """Generic request method for bulk upload of entities from a excel file. 
        Load the file from the request and if the records are valid, save them to the database, else return the errors in a excel file."""
        # Creating savepoint to rollback in case of error
        sid = transaction.savepoint()
        try:
            try:
                # Content type must be xlsm content type
                content_type = 'application/vnd.ms-excel.sheet.macroenabled.12'
                # Getting file from request
                template = request.FILES.get('template', None)
                # Checking if file is not empty
                if template is None:
                    raise serializers.ValidationError(
                        {'template': 'Este campo es requerido.'})
                # Checking if file is an excel file
                if template.content_type != content_type:
                    raise serializers.ValidationError(
                        {'template': 'Formato inválido de archivo. El formato permitido es xlsx.'})
                # Loading file on memory
                wb = openpyxl.load_workbook(
                    template, read_only=False, keep_vba=True)
                # Getting first worksheet
                ws = wb.worksheets[0]
                # Clearing all previous comments in worksheet
                self.clear_worksheet_styles_and_comments(ws)
                serializer_class = self.get_serializer_class()
                data = []
                has_errors = False
                field_names = ()
                for row in ws.rows:
                    # Getting tuple of cell values from row
                    row_values = tuple(c.value for c in row)
                    if not field_names:
                        # Getting field_names from first row
                        field_names = tuple(
                            f for f in row_values if f is not None)
                        continue
                    # Creating dict from row values
                    data = dict(zip(field_names, row_values))
                    # Removing None values
                    data = {k: v for k, v in data.items() if v is not None}
                    # Getting just values from data
                    values = list(data.values())
                    # Checking if not a empty row
                    if values.count(None) == len(values):
                        continue
                    serializer = serializer_class(data=data)
                    try:
                        serializer.is_valid(raise_exception=True)
                        serializer.save()
                    except serializers.ValidationError as e:
                        if not has_errors:
                            has_errors = True
                        # Indexing cells with field names
                        index_row = dict(zip(field_names, row))
                        for field_name, cell in index_row.items():
                            # Checking if field name is in error
                            if field_name in e.detail.keys():
                                # Getting error message
                                error_message = ', '.join(
                                    [str(ed) for ed in e.detail[field_name]])
                                # Creating comment with error message in cell
                                cell.comment = openpyxl.comments.Comment(
                                    error_message, 'System')
                                # Coloring cell background
                                cell.fill = openpyxl.styles.PatternFill(
                                    'solid', start_color='FF9490')
                # If there are errors, returning response template with errors
                if has_errors:
                    # Saving workbook
                    wb = openpyxl.writer.excel.save_virtual_workbook(wb)
                    response = HttpResponse(
                        wb, content_type=content_type, status=status.HTTP_400_BAD_REQUEST)
                    response['Content-Disposition'] = f'attachment; filename="{self.bulk_upload_template_name}"'
                    response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
                    # Rolling back savepoint
                    transaction.savepoint_rollback(sid)
                    return response
                response = Response({'detail': _('Successfully uploaded.')})
                # Commiting savepoint
                transaction.savepoint_commit(sid)
                return response
            except serializers.ValidationError as e:
                raise e
            except ValidationError as e:
                raise serializers.ValidationError({'template': e.message})
            except Exception as e:
                raise serializers.ValidationError(
                    {'detail': 'Algo ocurrió mal.'})
        except Exception as e:
            # Rolling back savepoint
            transaction.savepoint_rollback(sid)
            raise e


class BulkUploadStaffViewSet(BulkUploadViewSet):
    model = Staff
    queryset = Staff.objects.all()
    serializer_class = StaffListSerializer

class RessourceListViewSet(viewsets.ModelViewSet):
    model = Ressource
    queryset = Ressource.objects.all().order_by('name')
    serializer_class = RessourceListSerializer

class RessourceViewSet(viewsets.ModelViewSet):
    model = Ressource
    queryset = Ressource.objects.all()
    serializer_class = RessourceSerializer 

class R_SOViewSet(viewsets.ModelViewSet):
    model = R_SO
    queryset = R_SO.objects.all()
    serializer_class = R_SOSerializer
    filter_backends = [R_SOFilterBackend]

class RessourceWithServiceOfferedViewSet(viewsets.ModelViewSet):
    model = Ressource
    queryset = Ressource.objects.all()
    serializer_class = RessourceWithServiceOfferedSerializer


class GenerateServiceOffered(APIView):
    queryset = ServiceOffered.objects.all()

    def get(self,request,*args,**kwargs):
        services = ServiceOffered.objects.all()
        serviceOffered = []
        print(1)
        for s in services:
            serviceOffered.append({
                "name": s.name,
                "type": dict(ServiceOffered.TYPE).get(s.type),
                "area": s.area.name,
                "criticality": s.criticality,
                "max_scale": s.scale.max_value,
                "rto":s.maximum_recovery_time
            })
        
        data = {
            "serviceOffered": serviceOffered
        }

        pdf = render_to_pdf("service_offered.html",data)
        return HttpResponse(pdf,content_type='application/pdf')