import os

import openpyxl
from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.translation import ugettext_lazy as _
from django.views.static import serve
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

class BulkUploadViewSet(viewsets.ModelViewSet):
    bulk_upload_template_name = None

    @action(methods=['GET'], detail=False)
    def download_bulk_upload_template(self, request):
        template = serve(request, os.path.join('bulk_upload_templates',
                         self.bulk_upload_template_name), document_root=settings.STATIC_ROOT)
        template.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
        template.headers['Cache-Control'] = 'no-cache'
        return template

    @staticmethod
    def clear_worksheet_styles_and_comments(worksheet: Worksheet):
        is_first_row = True
        for row in worksheet.rows:
            if is_first_row:
                is_first_row = False
                continue
            for cell in row:
                # Removing comments and background colors
                cell.comment = None
                cell.fill = openpyxl.styles.PatternFill(fill_type=None)

    @action(methods=['POST'], detail=False)
    @transaction.non_atomic_requests()
    def bulk_upload(self, request):
        """Generic request method for bulk upload of entities from a excel file. 
        Load the file from the request and if the records are valid, save them to the database, else return the errors in a excel file."""
        # Creating savepoint to rollback in case of error
        sid = transaction.savepoint()
        try:
            try:
                # Content type must be xlsm content type
                content_type = 'application/vnd.ms-excel.sheet.macroenabled.12'
                # Getting file from request
                template = request.FILES.get('template', None)
                # Checking if file is not empty
                if template is None:
                    raise serializers.ValidationError(
                        {'template': _('This field is required.')})
                # Checking if file is an excel file
                if template.content_type != content_type:
                    raise serializers.ValidationError(
                        {'template': _('Invalid file format. The allowed format is xlsx.')})
                # Loading file on memory
                wb = openpyxl.load_workbook(
                    template, read_only=False, keep_vba=True)
                # Getting first worksheet
                ws = wb.worksheets[0]
                # Clearing all previous comments in worksheet
                self.clear_worksheet_styles_and_comments(ws)
                serializer_class = self.get_serializer_class()
                data = []
                has_errors = False
                field_names = ()
                for row in ws.rows:
                    # Getting tuple of cell values from row
                    row_values = tuple(c.value for c in row)
                    if not field_names:
                        # Getting field_names from first row
                        field_names = tuple(
                            f for f in row_values if f is not None)
                        continue
                    # Creating dict from row values
                    data = dict(zip(field_names, row_values))
                    # Removing None values
                    data = {k: v for k, v in data.items() if v is not None}
                    # Getting just values from data
                    values = list(data.values())
                    # Checking if not a empty row
                    if values.count(None) == len(values):
                        continue
                    serializer = serializer_class(data=data)
                    try:
                        serializer.is_valid(raise_exception=True)
                        serializer.save()
                    except serializers.ValidationError as e:
                        if not has_errors:
                            has_errors = True
                        # Indexing cells with field names
                        index_row = dict(zip(field_names, row))
                        for field_name, cell in index_row.items():
                            # Checking if field name is in error
                            if field_name in e.detail.keys():
                                # Getting error message
                                error_message = ', '.join(
                                    [str(ed) for ed in e.detail[field_name]])
                                # Creating comment with error message in cell
                                cell.comment = openpyxl.comments.Comment(
                                    error_message, 'System')
                                # Coloring cell background
                                cell.fill = openpyxl.styles.PatternFill(
                                    'solid', start_color='FF9490')
                # If there are errors, returning response template with errors
                if has_errors:
                    # Saving workbook
                    wb = openpyxl.writer.excel.save_virtual_workbook(wb)
                    response = HttpResponse(
                        wb, content_type=content_type, status=status.HTTP_400_BAD_REQUEST)
                    response['Content-Disposition'] = f'attachment; filename="{self.bulk_upload_template_name}"'
                    response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition'
                    # Rolling back savepoint
                    transaction.savepoint_rollback(sid)
                    return response
                response = Response({'detail': _('Successfully uploaded.')})
                # Commiting savepoint
                transaction.savepoint_commit(sid)
                return response
            except serializers.ValidationError as e:
                raise e
            except ValidationError as e:
                raise serializers.ValidationError({'template': e.message})
            except Exception as e:
                raise serializers.ValidationError(
                    {'detail': _('Something went wrong.')})
        except Exception as e:
            # Rolling back savepoint
            transaction.savepoint_rollback(sid)
            raise e